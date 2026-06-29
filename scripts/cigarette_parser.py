#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
卷烟货源投放策略表解析器（可复用 CLI 模块）

用法：
    python cigarette_parser.py --input <xlsx路径> --output-dir <输出目录>
    python cigarette_parser.py --input data/xxx.xlsx --output-dir output/

功能：
    解析 xlsx 中的三个工作表（档级投放、标签投放、雪茄投放），
    自动从文件内容提取投放时间，生成统一 JSON 文件。

输出文件名格式：卷烟投放策略_<投放时间>.json
    投放时间从 xlsx 单元格 A2 自动提取并清洗。
"""
import argparse
import json
import os
import re
import sys
import traceback

import openpyxl


# ============================================================
# 工具函数
# ============================================================
def extract_date_range(ws):
    """从工作表 A2 单元格提取投放时间区间。
    A2 形如：'投放时间：2026年6月21日下午-2026年6月26日上午  单位：条'
    提取为：'2026年6月21日-6月26日'
    """
    a2 = ws.cell(row=2, column=1).value or ''
    a2 = str(a2)
    # 匹配 "投放时间：xxx-xxx"
    m = re.search(r'投放时间[：:]\s*(.+?)(?:\s{2,}|\s+单位)', a2)
    if not m:
        # 退化：取"投放时间："之后到行尾或"单位"之前
        m = re.search(r'投放时间[：:]\s*(.+)', a2)
    raw = m.group(1).strip() if m else '未知时间'
    # 清洗：去掉"下午/上午"等时段词，简化为 日期-日期
    # 例：'2026年6月21日下午-2026年6月26日上午' -> '2026年6月21日-6月26日'
    cleaned = re.sub(r'[上下]午', '', raw)
    # 若后半段重复了年月，则只保留日
    parts = cleaned.split('-')
    if len(parts) == 2:
        left, right = parts[0].strip(), parts[1].strip()
        # 提取 left 的 年/月
        ym = re.match(r'(\d+年\d+月)', left)
        if ym and right.startswith(ym.group(1)):
            right = right[len(ym.group(1)):]
        cleaned = f'{left}-{right}'
    # 去掉文件名非法字符
    cleaned = re.sub(r'[\\/:*?"<>|]', '', cleaned)
    return cleaned


def safe_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ============================================================
# 第一部分：档级投放
# ============================================================
def parse_grade_sheet(wb):
    ws = wb['档级投放']
    # 列：A=1(代码) B=2(名称) C=3(品类) D=4(区域) E=5(投放户数) F..AI=6..35(30档..1档) AJ=36(备注)
    grade_col = {grade: 6 + (30 - grade) for grade in range(1, 31)}

    result = {"档位汇总": {}, "条件投放策略": []}
    for grade in range(1, 31):
        result["档位汇总"][f"{grade}档"] = {"卷烟数量": 0, "卷烟列表": []}

    def find_related_cigarette(r):
        for rr in range(r - 1, 4, -1):
            c = ws.cell(row=rr, column=1).value
            n = ws.cell(row=rr, column=2).value
            if c is not None and n is not None:
                if isinstance(c, str) and ('制表' in c or '日期' in c):
                    return None, None, None
                return safe_str(c), safe_str(n), safe_str(ws.cell(row=rr, column=3).value)
        return None, None, None

    for r in range(5, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        cat = ws.cell(row=r, column=3).value
        region = ws.cell(row=r, column=4).value
        households = ws.cell(row=r, column=5).value
        f_val = ws.cell(row=r, column=6).value
        note = ws.cell(row=r, column=36).value

        if isinstance(code, str) and ('制表' in code or '日期' in code):
            continue

        is_strategy_row = (code is None and name is None
                           and isinstance(f_val, str) and len(f_val.strip()) > 5)
        if is_strategy_row:
            rel_code, rel_name, rel_cat = find_related_cigarette(r)
            result["条件投放策略"].append({
                "关联卷烟代码": rel_code,
                "关联卷烟名称": rel_name,
                "关联卷烟品类": rel_cat,
                "投放策略说明": f_val.strip(),
                "投放户数": households,
                "区域": safe_str(region),
                "所属品类(行内)": safe_str(cat)
            })
            continue

        if code is None or name is None:
            continue

        name = str(name).strip()
        code = str(code).strip()

        for grade in range(1, 31):
            col = grade_col[grade]
            v = ws.cell(row=r, column=col).value
            if is_num(v) and v > 0:
                result["档位汇总"][f"{grade}档"]["卷烟列表"].append({
                    "卷烟代码": code,
                    "卷烟名称": name,
                    "投放数量(条)": v,
                    "投放户数": households,
                    "区域": safe_str(region),
                    "备注": safe_str(note)
                })

    for grade in range(1, 31):
        items = result["档位汇总"][f"{grade}档"]["卷烟列表"]
        result["档位汇总"][f"{grade}档"]["卷烟数量"] = len(items)
        items.sort(key=lambda x: (-x["投放数量(条)"], x["卷烟名称"]))

    return result


# ============================================================
# 第二部分：标签投放
# ============================================================
def parse_label_sheet(wb):
    ws = wb['标签投放']
    result = []
    r = 1
    while r <= ws.max_row:
        a_val = ws.cell(row=r, column=1).value
        is_title = False
        if isinstance(a_val, str) and a_val.strip():
            s = a_val.strip()
            if '-' in s:
                prefix, rest = s.split('-', 1)
                if prefix.strip().isdigit() and rest.strip() and not rest.strip().endswith('档'):
                    is_title = True
        if is_title:
            strategy_title = a_val.strip()
            num = strategy_title.split('-')[0].strip()
            desc = strategy_title[len(num) + 1:].strip()
            header_row = r + 1
            cig_names = []
            for c in range(3, ws.max_column + 1):
                v = ws.cell(row=header_row, column=c).value
                if v is not None and str(v).strip() and str(v).strip() != '合计':
                    cig_names.append((c, str(v).strip()))
            a_header = ws.cell(row=header_row, column=1).value
            has_grade = (a_header is not None and '档级' in str(a_header))
            order_col = 2 if has_grade else 1

            data_rows = []
            dr = header_row + 1
            last_grade = None
            while dr <= ws.max_row:
                da = ws.cell(row=dr, column=1).value
                db = ws.cell(row=dr, column=2).value
                if (da is None or str(da).strip() == '') and (db is None or str(db).strip() == ''):
                    break
                if isinstance(da, str) and da.strip() and '-' in da:
                    s = da.strip()
                    prefix, rest = s.split('-', 1)
                    if prefix.strip().isdigit() and rest.strip() and not rest.strip().endswith('档'):
                        break
                grade_val = safe_str(da)
                if has_grade:
                    if grade_val:
                        last_grade = grade_val
                    else:
                        grade_val = last_grade
                order_val_cell = ws.cell(row=dr, column=order_col).value
                order_val = safe_str(order_val_cell)
                row_data = {"档级": grade_val, "订购量": order_val, "卷烟投放": []}
                for col, cname in cig_names:
                    v = ws.cell(row=dr, column=col).value
                    if is_num(v) and v > 0:
                        row_data["卷烟投放"].append({"卷烟名称": cname, "投放数量(条)": v})
                data_rows.append(row_data)
                dr += 1

            cig_list = []
            for row_data in data_rows:
                condition = row_data["档级"] if has_grade else row_data["订购量"]
                for cd in row_data["卷烟投放"]:
                    cig_list.append({
                        "卷烟名称": cd["卷烟名称"],
                        "投放数量(条)": cd["投放数量(条)"],
                        "适用条件": condition,
                        "档级": row_data["档级"] if has_grade else None,
                        "订购量": row_data["订购量"]
                    })
            cig_list.sort(key=lambda x: (x["卷烟名称"], -x["投放数量(条)"]))

            result.append({
                "策略编号": num,
                "策略说明": desc,
                "卷烟数量": len(set(x["卷烟名称"] for x in cig_list)),
                "卷烟列表": cig_list
            })
            r = dr
        else:
            r += 1
    return result


# ============================================================
# 第三部分：雪茄投放
# ============================================================
def parse_cigar_sheet(wb):
    ws = wb['雪茄投放']
    # 列：A=1(类别) B=2(代码) C=3(名称) D=4(货源属性) E=5(区域) F=6(投放户数)
    #      G..M=7..13(A档..E-档) N=14(标签投放) O=15(备注)
    grade_cols = [(7, 'A档'), (8, 'B档'), (9, 'C档'), (10, 'D档'),
                  (11, 'E+档'), (12, 'E档'), (13, 'E-档')]

    result = {"档位投放": {g[1]: [] for g in grade_cols}, "标签投放": []}

    # 类别向下填充（合并单元格）
    last_category = None
    for r in range(5, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        if cat is not None and '制表' not in str(cat) and '本周策略' not in str(cat):
            last_category = safe_str(cat)
        code = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=3).value
        if code is None or name is None:
            continue
        if isinstance(code, str) and ('制表' in code or '日期' in code):
            continue

        attr = safe_str(ws.cell(row=r, column=4).value)
        region = safe_str(ws.cell(row=r, column=5).value)
        households = ws.cell(row=r, column=6).value
        label = safe_str(ws.cell(row=r, column=14).value)
        note = safe_str(ws.cell(row=r, column=15).value)

        base = {
            "卷烟代码": safe_str(code),
            "卷烟名称": safe_str(name),
            "类别": last_category,
            "货源属性": attr,
            "投放户数": households,
            "区域": region,
            "备注": note
        }

        # 档位投放
        for col, gname in grade_cols:
            v = ws.cell(row=r, column=col).value
            if is_num(v) and v > 0:
                item = dict(base)
                item["投放档位"] = gname
                item["投放数量(条)"] = v
                result["档位投放"][gname].append(item)

        # 标签投放
        if label:
            item = dict(base)
            item["标签投放说明"] = label
            result["标签投放"].append(item)

    for gname in result["档位投放"]:
        result["档位投放"][gname].sort(key=lambda x: (-x["投放数量(条)"], x["卷烟名称"]))

    return result


# ============================================================
# 主流程
# ============================================================
def parse_file(input_path):
    """解析单个 xlsx 文件，返回 (date_range, result_dict)。"""
    wb = openpyxl.load_workbook(input_path, data_only=True)

    # 校验必需的工作表
    required = ['档级投放', '标签投放', '雪茄投放']
    missing = [s for s in required if s not in wb.sheetnames]
    if missing:
        raise ValueError(f'缺少工作表: {missing}，现有: {wb.sheetnames}')

    date_range = extract_date_range(wb['档级投放'])

    grade_data = parse_grade_sheet(wb)
    label_data = parse_label_sheet(wb)
    cigar_data = parse_cigar_sheet(wb)

    result = {
        "投放时间": date_range,
        "单位": "条",
        "数据来源": os.path.basename(input_path),
        "档级投放": grade_data,
        "标签投放": label_data,
        "雪茄投放": cigar_data
    }
    return date_range, result


def main():
    parser = argparse.ArgumentParser(description='卷烟货源投放策略表解析器')
    parser.add_argument('--input', required=True, help='输入 xlsx 文件路径')
    parser.add_argument('--output-dir', default='output', help='输出目录（默认 output）')
    parser.add_argument('--output', default=None, help='输出文件完整路径（覆盖 --output-dir）')
    args = parser.parse_args()

    input_path = args.input
    if not os.path.exists(input_path):
        print(f'错误：文件不存在 {input_path}', file=sys.stderr)
        sys.exit(1)

    print(f'解析中: {input_path}')
    try:
        date_range, result = parse_file(input_path)
    except Exception as e:
        print(f'解析失败: {e}', file=sys.stderr)
        traceback.print_exc()
        sys.exit(2)

    # 确定输出路径
    if args.output:
        out_path = args.output
    else:
        os.makedirs(args.output_dir, exist_ok=True)
        out_path = os.path.join(args.output_dir, f'卷烟投放策略_{date_range}.json')

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f'已生成: {out_path}')
    # 打印简要汇总
    g = result['档级投放']
    total_grade = sum(v['卷烟数量'] for v in g['档位汇总'].values())
    print(f'  [档级投放] 档位记录 {total_grade} 条, 条件投放策略 {len(g["条件投放策略"])} 条')
    print(f'  [标签投放] 策略 {len(result["标签投放"])} 个')
    c = result['雪茄投放']
    total_cigar = sum(len(v) for v in c['档位投放'].values())
    print(f'  [雪茄投放] 档位记录 {total_cigar} 条, 标签投放 {len(c["标签投放"])} 条')


if __name__ == '__main__':
    main()
